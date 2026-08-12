# -*- coding: utf-8 -*-
"""Re-theme docs/Prime_Capital_Bank_Analytics.xlsx to the standing PCB palette.

The workbook previously used a close-but-not-identical navy/cyan (#1A2E3B /
#00A8E0). This aligns it exactly to the site's palette so every deliverable -
web pages and the Excel workbook alike - uses the same colours:

  navy   #0A1628  title bars
  navy2  #0F2040  column header row
  gold   #C9A84C  header row text, positive/primary accents
  cyan   #00B4CC  reserved for KPI highlight cells only
  text   #1C2B3A  body text
  dim    #6B7E92  secondary/muted text
  border #DDE5F0  gridlines

Preserves every value and formula - only fill/font colours change.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

PATH = r"C:\Users\Anthony.DESKTOP-ES5HL78\Documents\prime-capital-bank\docs\Prime_Capital_Bank_Analytics.xlsx"

NAVY, NAVY2, GOLD, CYAN = "0A1628", "0F2040", "C9A84C", "00B4CC"
TEXT, DIM, BORDER, WHITE = "1C2B3A", "6B7E92", "DDE5F0", "FFFFFF"

fill_navy  = PatternFill("solid", fgColor=NAVY)
fill_navy2 = PatternFill("solid", fgColor=NAVY2)
fill_white = PatternFill("solid", fgColor=WHITE)
font_title = Font(color=WHITE, bold=True, size=13, name="Segoe UI")
font_hdr   = Font(color=GOLD, bold=True, size=10, name="Segoe UI")
font_body  = Font(color=TEXT, bold=False, size=10, name="Segoe UI")
font_label = Font(color=TEXT, bold=True, size=10, name="Segoe UI")
thin = Side(style="thin", color=BORDER)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.load_workbook(PATH)
changed = 0

for name in wb.sheetnames:
    ws = wb[name]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
            fill = fill[-6:] if fill and fill != "00000000" else None
            was_bold = bool(cell.font and cell.font.bold)

            if cell.row == 1:
                # workbook title bar
                cell.fill = fill_navy
                cell.font = font_title
            elif fill and fill.upper() in ("1A2E3B",):
                cell.fill = fill_navy
                cell.font = font_title
            elif fill and fill.upper() in ("00A8E0",):
                # column header row -> navy2 fill, gold text (was cyan fill/white text)
                cell.fill = fill_navy2
                cell.font = font_hdr
            elif fill and fill.upper() == "FFFFFF":
                cell.fill = fill_white
                cell.font = font_label if was_bold else font_body
                cell.border = border_all
            changed += 1

wb.save(PATH)
print(f"Re-themed {len(wb.sheetnames)} sheets, {changed} styled cells touched.")
print("Sheets:", wb.sheetnames)
